import os
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font

# Base directory
base_dir = Path(r'c:\repos\powerbi-docs-pr\powerbi-docs')
includes_dir = base_dir / 'includes'

# Find all INCLUDE files
include_files = list(includes_dir.glob('**/*.md'))
print(f"Found {len(include_files)} INCLUDE files")

# Prepare data
data = []

for include_file in include_files:
    # Get relative path from powerbi-docs directory
    rel_path = include_file.relative_to(base_dir)
    
    # Extract ms.date from the file
    ms_date = None
    try:
        with open(include_file, 'r', encoding='utf-8') as f:
            content = f.read()
            # Look for ms.date in the front matter
            match = re.search(r'ms\.date:\s*([^\n]+)', content)
            if match:
                ms_date = match.group(1).strip()
    except Exception as e:
        print(f"Error reading {rel_path}: {e}")
    
    # Count references to this INCLUDE file
    # INCLUDE files are referenced using [!INCLUDE [filename](path/to/file.md)]
    include_name = include_file.name
    include_rel_path = str(rel_path).replace('\\', '/')
    
    reference_count = 0
    
    # Search all .md files in the repo
    for md_file in base_dir.glob('**/*.md'):
        # Skip the INCLUDE file itself
        if md_file == include_file:
            continue
        
        try:
            with open(md_file, 'r', encoding='utf-8') as f:
                md_content = f.read()
                # Look for references to this INCLUDE file
                # Pattern: [!INCLUDE [something](../includes/filename.md)]
                # The path can be relative, so we'll search for the filename
                pattern = r'\[!INCLUDE\s*\[.*?\]\(.*?' + re.escape(include_name) + r'\)\]'
                matches = re.findall(pattern, md_content, re.IGNORECASE)
                reference_count += len(matches)
        except Exception as e:
            # Skip files that can't be read
            pass
    
    data.append({
        'path': str(rel_path).replace('\\', '/'),
        'ms_date': ms_date or '',
        'references': reference_count
    })
    
    print(f"Processed: {rel_path} - {reference_count} references")

# Sort by path
data.sort(key=lambda x: x['path'])

# Create Excel workbook
wb = Workbook()
ws = wb.active
ws.title = "All INCLUDE Files"

# Add headers
headers = ['File Path', 'ms.date', 'Reference Count']
ws.append(headers)

# Style headers
for cell in ws[1]:
    cell.font = Font(bold=True)

# Add data
for item in data:
    ws.append([item['path'], item['ms_date'], item['references']])

# Adjust column widths
ws.column_dimensions['A'].width = 60
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 18

# Create second sheet for unreferenced files
ws2 = wb.create_sheet("Unreferenced (0 refs)")
ws2.append(headers)

# Style headers
for cell in ws2[1]:
    cell.font = Font(bold=True)

# Add unreferenced files
unreferenced_count = 0
for item in data:
    if item['references'] == 0:
        ws2.append([item['path'], item['ms_date'], item['references']])
        unreferenced_count += 1

# Adjust column widths
ws2.column_dimensions['A'].width = 60
ws2.column_dimensions['B'].width = 15
ws2.column_dimensions['C'].width = 18

# Save workbook
output_file = r'c:\repos\powerbi-docs-pr\include_files_analysis.xlsx'
wb.save(output_file)
print(f"\nExcel file created: {output_file}")
print(f"Total INCLUDE files analyzed: {len(data)}")
print(f"Unreferenced INCLUDE files (0 references): {unreferenced_count}")
print(f"Referenced INCLUDE files: {len(data) - unreferenced_count}")
